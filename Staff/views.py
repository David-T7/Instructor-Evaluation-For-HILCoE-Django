from django.shortcuts import redirect, render
from Course.models import Course, CourseInstructor, Term
from Evaluation.models import Criteria, EvaluationCriteria
from Instructor.models import Instructor
from Student.models import StudentCourseEnrollment, StudentEvaluationResult
from Student.forms import GeneralReportForm
from .models import Staff, StaffEvaluationResult
from django.contrib import  messages
from django.utils import timezone
from django.shortcuts import render
from io import BytesIO
from django.http import HttpResponse
from django.template.loader import get_template
from django.views import View
from xhtml2pdf import pisa
from django.http import HttpResponse
from openpyxl import Workbook
from Student.views import total_evaluation_reports_from_query
# Create your views here.
def stafflandingpage(request):
    # for displaying a staff page 
    return render(request , 'stafflandingpage.html')
def staffHomePage(request):
    # for dispalying a staff home page
    term = None
    try:
        term = Term.objects.get(EvaluationDone = False)
    except:
        term = None
    evaluation_started = False
    evaluation_ended = False
    if(term.Evaluation_Start_Date <= timezone.now()): #check if evaluation started
        print('evaluation started')
        evaluation_started = True
    if(term.Evaluation_End_Date < timezone.now()): #check if evaluation ended
        print('evaluation ended')
        evaluation_ended = True
    context = { 'active_page': 'home', 'term':term , 'evaluation_started':evaluation_started , 'evaluation_ended':evaluation_ended }
    return render(request , 'staff/staffhome.html' , context)

def staff_evaluate_page(request):
    # for displaying currently courses given by instructors
    term = None
    try:
        term = Term.objects.get(EvaluationDone = False)
    except:
        term = None
    evaluation_started = False
    evaluation_ended = False
    if(term.Evaluation_Start_Date <= timezone.now()): #check if evaluation started
        print('evaluation started')
        evaluation_started = True
    if(term.Evaluation_End_Date < timezone.now()): #check if evaluation ended
        print('evaluation ended')
        evaluation_ended = True
    staff = Staff.objects.get(Account_id=request.user)
    student_enrollments = StudentCourseEnrollment.objects.filter(term=term)
    courses = []
    staff_evaluation_result = StaffEvaluationResult.objects.filter(Staff_id = staff)
    evaluated_courses_types = []
    evaluated_instructors = {}
    course_instructors = []    
    if staff_evaluation_result:
        for staff_evaluation in staff_evaluation_result:
            eval_list = []
            eval_list.append(staff_evaluation.Instructor_id.Instructor_id)
            eval_list.append(staff_evaluation.CourseType)
            eval_list.append(staff_evaluation.Course_id.CourseName)
            evaluated_instructors[staff_evaluation.Course_id.CourseName+staff_evaluation.CourseType] = eval_list
    courses_data = []
    for enrollment in student_enrollments:
        if enrollment.course not in courses:
            courses.append(enrollment.course)
            instructors_data = []
            for courseinstructors in CourseInstructor.objects.filter(Course = enrollment.course ):
                course_instructors.append(courseinstructors)
                instructor_info = {
                'name': f'{courseinstructors.Instructors.FirstName} {courseinstructors.Instructors.LastName}',
                'title': courseinstructors.Instructors.Title,
                'instructor_id':courseinstructors.Instructors.Instructor_id , 
                'profile_pic': courseinstructors.Instructors.ProfilePic.url if courseinstructors.Instructors.ProfilePic else None,
                }

            course_data = {
            'course_name': enrollment.course.CourseName,
            'course_id':enrollment.course.Course_id,
            'instructors': instructors_data,
            'evaluate_url': f'/evaluate/{enrollment.id}/',  # Replace with the actual URL for evaluation
            }
            courses_data.append(course_data)

    context = {
        'courses_data': courses_data,
        'term': term,
        'active_page':'evaluation',
        'staff':staff,
        'evaluated_courses_types':evaluated_courses_types,
        'course_instructors':course_instructors,
        'evaluated_instructors':evaluated_instructors,
        'evaluation_started':evaluation_started,
        'evaluation_ended':evaluation_ended,
        
    }

    return render(request, 'staff/evaluate.html', context)

def staff_evaluate_course(request, staff_id, course_id, instructor_id  , course_type):
    # for displaying the evaluation form and saving the results of the evaluation
    evaluationMapping= {'Excellent':5,'Very Good':4,'Good':3,'Poor':2,'Very Poor':1}
    evaluated_criteria_descriptions = set()
    staff = Staff.objects.get(Account_id=request.user)
    course = Course.objects.get(Course_id=course_id)
    instructor = Instructor.objects.get(Instructor_id=instructor_id)
    # Find the latest term where evaluation is not done
    term = Term.objects.filter(Courses_Given=course, EvaluationDone=False).order_by('-Year', 'Season').first()
    desired_order = ['Timely Grade Submission' , 'Accuracy of Grade Records' , 'Grade Appeal Process' , 'Grade Change Procedures']
    if not term:
        messages.warning(request, 'Course evaluation is already completed for all terms.')
        return redirect('evaluate')  # Redirect to home or another page
    all_criteria_objects = EvaluationCriteria.objects.get(Evaluator='StaffMember')  # Adjust based on your criteria
    # all_criteria_objects = EvaluationCriteria.objects.all()
    # Collect all distinct criteria sections using Python
    all_criteria_data = all_criteria_objects.Criteria_data.all()
    all_criteria = []
    criteria_sections = []
    for criteria_object in all_criteria_data:
        criteria = Criteria.objects.get(Criteria_id= criteria_object.Criteria_id) 
        if criteria.Section not in  criteria_sections :
            all_criteria.append(criteria)
            criteria_sections.append(criteria.Section)
        print("criteria section" , criteria.Section )
    print("all sectins are ", all_criteria)
    criteria_sections.sort(key=lambda x: desired_order.index(x.Section))
    criteria_results = {}
    error_occured = False
    if request.method == 'POST':
        submitted_data = request.POST.copy()
        for key, value in submitted_data.items():
                if key.startswith('criteria_'):
                    criteria_id = key.replace('criteria_', '')
                    criteria_results[criteria_id] = value
        evaluation_result = {}
        for criteria in all_criteria_data:
            score = request.POST.get(f'criteria_{criteria.Criteria_id}')
            if score:
                section_name = criteria.Section.Section
                if section_name not in evaluation_result:
                        evaluation_result[section_name] = {}
                evaluation_result[section_name][str(criteria.description)] = evaluationMapping[score]
                evaluated_criteria_descriptions.add(criteria.description)
        # Get all unique criteria descriptions
        all_criteria_descriptions = set(criteria.description for criteria in all_criteria_data)        
        print("evaluation resullt",evaluation_result)
        additional_comment = request.POST.get('additional_comment')     
        if evaluation_result and evaluated_criteria_descriptions == all_criteria_descriptions:
            result_instance, created = StaffEvaluationResult.objects.get_or_create(
                Staff_id=staff,
                Course_id=course,
                Instructor_id=instructor,
                CourseType = course_type,
                Term_id=term,
                AdditionalComment = additional_comment, 
                defaults={'EvaluationResult': evaluation_result, 'EvaluationDone': True}
            )

            if not created:
                messages.error('Please evaluate all criteria descriptions before submitting!')
                error_occured = True
            else:
                messages.success(request, 'You have completed Evaluation for course '+ course.CourseName) 
                return redirect('evaluate_staff')  # Redirect to home or another page after evaluation
        else:
            messages.error(request, 'Please evaluate all criteria descriptions before submitting!')
            error_occured = True

    context = {
        'course':course,
        'criteria_data': all_criteria_data, 
        'active_page':'evaluation',
        'all_criteria_sections':all_criteria ,
        'criteria_sections':criteria_sections,
        'instructor':instructor,
        'course_type':course_type,
        'criteria_results' : criteria_results,
        'error_occured'  :error_occured,
    }

    return render(request, 'staff/evaluate_course.html', context)

def academicheadHomePage(request):
    # for displaying the academic head homepage
    term = None
    try:
        term = Term.objects.get(EvaluationDone = False)
    except:
        term = None
    evaluation_started = False
    evaluation_ended = False
    if(term.Evaluation_Start_Date <= timezone.now()): #check if evaluation started
        print('evaluation started')
        evaluation_started = True
    if(term.Evaluation_End_Date < timezone.now()): #check if evaluation ended
        print('evaluation ended')
        evaluation_ended = True
    context = { 'active_page': 'home', 'term':term , 'evaluation_started':evaluation_started , 'evaluation_ended':evaluation_ended}
    return render(request , 'academichead/academicheadhome.html' , context)


def ViewPDF(templatename , data):
    # for displaying html content as pdf
	pdf = render_to_pdf('academichead/'+templatename, data)
	return HttpResponse(pdf, content_type='application/pdf')

def render_to_pdf(template_path, context_dict):
    # for rendering the content on the given html page
    template = get_template(template_path)
    html = template.render(context_dict)
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)
    
    if not pdf.err:
        return HttpResponse(result.getvalue(), content_type='application/pdf')
    return None

def DownloadPDF(templatename, data, filename):
    # for downloading the content on the given html page as pdf
    pdf = render_to_pdf('academichead/'+templatename, data)

    if pdf:
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'
        return response

    # Handle the case where PDF generation failed
    return HttpResponse("PDF generation failed", status=500)

def generalEvaluationReport(request , type):
    # for genrating evaluation report basd on a given input
    evaluations = None
    term = None
    evaluator = None
    department = None
    course_type = None
    
    if request.method == 'POST':
        print("in post before evaluting form" , request.POST)
        query = request.POST
        query_dic = {'term':'Term_id'}
        if query:
            query_params = {}
            # query_dic = {'course':'Course_id' , 'term':'Term_id' , 'instructor':'Instructor_id'}
            if query['evaluator']:
                evaluator = query['evaluator']
            else:
                evaluator = 'both'
            if query['course_type']:
                course_type = query['course_type']
            if query['department']:
                department = query['department']
            print('form cleaned data is ' , query.items())
            # Check each form field and add it to the query_params if it's not empty
            for field_name, value in query.items():
                if value and field_name in query_dic:
                    if field_name == 'term':
                        value_object = Term.objects.get(Season = value.split()[0] ,Year=value.split()[1]) 
                        if value_object:
                            term = value_object    
  
            desired_order = []
            
            if evaluator == 'Student':
                evaluations = StudentEvaluationResult.objects.filter(
                Term_id= term , 
                EvaluationDone=True ,
                CourseType = course_type
                    )
                desired_order = ['Course Organization', 'Knowledge of the subject matter', 'Teaching Methods',
                    'Student Involvement', 'Evaluation Methods' , 'Personality Traits' , 'Availability and Support']
            elif evaluator == 'Staff':
                evaluations = StaffEvaluationResult.objects.filter(
                Term_id=term,
                EvaluationDone=True ,
                CourseType = course_type
                    )
                desired_order = ['Timely Grade Submission' , 'Accuracy of Grade Records' , 'Grade Appeal Process' , 'Grade Change Procedures']
            
            criteria_average_details = []
            criteria = []
            criteria_category = {}
            criteria_sections = []
            criteria_average_Scores = []
            courses = []
            for evaluation in evaluations:
                if evaluation.Course_id not in courses and evaluation.Course_id.Department == department:
                    courses.append(evaluation.Course_id) 
            if evaluations:
                for evaluation in evaluations:
                    if ( evaluation.Course_id in courses):
                        for category, sub_dict in evaluation.EvaluationResult.items():
                            if category not in criteria_sections:criteria_sections.append(category)
                            average_score = 0
                            for criterion, score in sub_dict.items():
                                if criterion not in criteria:
                                    criteria_category[criterion] = category
                                criteria.append(criterion)
                                criteria_average_details.append({
                                    'category': category,
                                    'criteria': criterion,
                                    'score': score,
                                    })
                print ('before soring criteria category is ' , criteria_category)
                criteria_sections  = sorted(criteria_sections, key=lambda x: desired_order.index(x) if x in desired_order else float('inf'))
                print ("criteria dic is " , criteria_category)
                for  criteria , category in criteria_category.items():
                    average_score = 0
                    len = 0
                    for average_details in criteria_average_details:
                        if average_details['category'] == category and average_details['criteria'] == criteria:
                            average_score += average_details['score']
                            len+=1
                    criteria_average_Scores.append(
                            {
                                'category': category,
                                'criteria': criteria,
                                'score': average_score / len,  
                             }
                                 )
                    print('criteria category before sorting is ' , criteria_category)
                    print ('criteria_average_Scores is ' , criteria_average_Scores)               
                context = {'criteria_average_Scores':criteria_average_Scores  , 
                        'criteria_sections':criteria_sections , 
                        'evaluations':evaluations, 
                         'active_page':'geneal_report',
                         'evaluator':evaluator ,
                         'term':term ,
                         'department':department,
                         'course_type':str.title(course_type),
                         'sender':'totalreport',
                    }
                if(type == 'view'):    
                    return render (ViewPDF('generalreport.html',context))
                elif (type == 'download'):
                    print('downloading')
                    return ViewPDF('generalreport.html' , context)
                    # return DownloadPDF('generalreport.html' , context , 'evaluation_report')
            else:
                messages.error(request, 'No result found!')
                terms = Term.objects.all()
                form = GeneralReportForm(initial={'course_type': query.get('course_type') , 
                                                  'evaluator': query.get('evaluator') ,
                                                  'term': query.get('term')})
                print('result not found ' , query)
                context = {'form':form , 'active_page': 'general_report' , 'terms':terms , 'query':query }
                return render(request, 'academichead/evaluationreportpage.html' , context )  
    else:
        terms = Term.objects.all()
        form = GeneralReportForm
        context = {'form':form , 'active_page': 'general_report' , 'terms':terms}
        return render(request, 'academichead/evaluationreportpage.html' , context )    


def generate_total_report_excel(request):
     # for genrating evaluation report basd on a given input
    evaluations = None
    term = None
    evaluator = None
    department = None
    course_type = None
    print("in generate excel file")
    if request.method == 'POST':
        print("in post before evaluting form" , request.POST)
        query = request.POST
        query_dic = {'term':'Term_id'}
        if query:
            query_params = {}
            for field_name, value in query.items():
                if value and field_name in query_dic:
                    if field_name == 'course':
                        try:
                            value_object = Course.objects.get(Course_id = value.split('(')[0])
                        except:
                            value_object = None
                    elif field_name == 'term':
                        try:
                            value_object = Term.objects.get(Season = value.split()[0] ,Year=value.split()[1])
                            if value_object:
                                term = value_object
                        except:
                            value_object = None
                    elif field_name == 'instructor':
                        try:
                            value_object = Instructor.objects.get(FirstName = value.split()[0] , LastName = value.split()[1] )
                        except:
                            value_object =None
                    query_params[query_dic[f"{field_name}"]] = value_object
                    # print("field name is",field_name , value)
            if query['evaluator']:
                evaluator = query['evaluator']
            else:
                evaluator = 'both'
            if query['course_type']:
                course_type = query['course_type']
            else:
                course_type = None
            if query['department']:
                department = query['department']
            else:
                department = None
            if evaluator == 'Total':
                    # print("evaluator is total before preceeding.....")
                    details = {"department":department , "term":term , "generate_excel_file":True , 'course_type':course_type}
                    return total_evaluation_reports_from_query(request , query_params ,details )
            print('form cleaned data is ' , query.items())
  
            desired_order = []
            
            if evaluator == 'Student':
                evaluations = StudentEvaluationResult.objects.filter(
                Term_id= term , 
                EvaluationDone=True ,
                CourseType = course_type
                    )
                desired_order = ['Course Organization', 'Knowledge of the subject matter', 'Teaching Methods',
                    'Student Involvement', 'Evaluation Methods' , 'Personality Traits' , 'Availability and Support']
            elif evaluator == 'Staff':
                evaluations = StaffEvaluationResult.objects.filter(
                Term_id=term,
                EvaluationDone=True ,
                CourseType = course_type
                    )
                desired_order = ['Timely Grade Submission' , 'Accuracy of Grade Records' , 'Grade Appeal Process' , 'Grade Change Procedures']

            # criteria_average_details = []
            # criteria = []
            # criteria_category = {}
            # criteria_sections = []
            # criteria_average_Scores = []
            # category_average_scores = []
            courses = []
            context = []
            course_instructors = CourseInstructor.objects.filter(CourseType = course_type)
            print("course instructors are ", course_instructors)
            # print("cousre instructor length is " , course_instructors.count)
            for evaluation in evaluations:
                if evaluation.Course_id not in courses and evaluation.Course_id.Department == department:
                    courses.append(evaluation.Course_id)
            if evaluations and courses:    
                for course_instructor in course_instructors:
                        criteria_average_details = []
                        criteria = []
                        criteria_category = {}
                        criteria_sections = []
                        criteria_average_Scores = []
                        category_average_scores = []
                        for evaluation in evaluations:
                            print("in evaluation")
                            if ( evaluation.Course_id in courses and str(evaluation.Instructor_id.Instructor_id) == str(course_instructor.Instructors.Instructor_id) and str(evaluation.Course_id.Course_id) == str(course_instructor.Course.Course_id) ):
                                print("passed now",course_instructor)
                                for category, sub_dict in evaluation.EvaluationResult.items():
                                    if category not in criteria_sections:criteria_sections.append(category)
                                    average_score = 0
                                    for criterion, score in sub_dict.items():
                                        if criterion not in criteria:
                                            criteria_category[criterion] = category
                                        criteria.append(criterion)
                                        criteria_average_details.append({
                                            'category': category,
                                            'criteria': criterion,
                                            'score': score,
                                            })
                        # print ('before soring criteria category is ' , criteria_category)
                        criteria_sections  = sorted(criteria_sections, key=lambda x: desired_order.index(x) if x in desired_order else float('inf'))
                        # print ("criteria dic is " , criteria_category)
                        print("category average details before iterations ", criteria_average_details)
                        for  criteria , category in criteria_category.items():
                            average_score = 0
                            len = 0
                            for average_details in criteria_average_details:
                                if average_details['category'] == category and average_details['criteria'] == criteria:
                                    average_score += average_details['score']
                                    len+=1
                            criteria_average_Scores.append(
                                    {
                                        'category': category,
                                        'criteria': criteria,
                                        'score': average_score / len,  
                                    }
                                        )
                            # Create a dictionary to store accumulated scores and counts for each category
                            # print('criteria category before sorting is ' , criteria_category)
                            # print ('criteria_average_Scores is ' , criteria_average_Scores)
                        category_averages = {}

                        # Iterate through the list and accumulate scores for each category
                        for entry in criteria_average_Scores:
                            category = entry['category']
                            score = entry['score']

                            if category not in category_averages:
                                category_averages[category] = {'total_score': 0, 'count': 0}

                            category_averages[category]['total_score'] += score
                            category_averages[category]['count'] += 1

                        # Calculate averages for each category
                        total_average_score = 0
                        len = 0
                        for category, data in category_averages.items():
                            average_score = data['total_score'] / data['count'] if data['count'] > 0 else 0
                            category_average_scores.append({category : average_score})
                            total_average_score += average_score
                            len +=1
                            print(f'Category: {category}, Average Score: {average_score}')
                        if (len > 0):
                            total_average_score /= len
                        # Accumulate the average score for the category
                        if(total_average_score > 0 ):
                            context.append({'category_average_scores':category_average_scores  , 
                                'title':course_instructor.Instructors.Title,
                                'first_name':course_instructor.Instructors.FirstName,
                                'last_name':course_instructor.Instructors.LastName , 
                                'course_name':course_instructor.Course.CourseName ,
                                'course_id':course_instructor.Course.Course_id , 
                                'total_score':round(total_average_score, 2),
                            } )
                details = {
                    'term':term,
                    'evaluator':evaluator,
                    'desired_order':desired_order,
                    'department':department,
                    'coursetype':course_type,
                }
                print("before generating excel file context file looks like.. ", context)
                print("courses before generating excel file " , courses)
                print("evaluation before generating excel file " , evaluations)
                
                return generate_excel(request , context, details )
            else:
                    messages.error(request, 'No result found!')
                    terms = Term.objects.all()
                    form = GeneralReportForm(initial={'course_type': query.get('course_type') , 
                                                    'evaluator': query.get('evaluator') ,
                                                    'term': query.get('term')})
                    print('result not found ' , query)
                    context = {'form':form , 'active_page': 'general_report' , 'terms':terms , 'query':query }
                    return render(request, 'academichead/evaluationreportpage.html' , context )  
    else:
        terms = Term.objects.all()
        form = GeneralReportForm
        context = {'form':form , 'active_page': 'general_report' , 'terms':terms}
        return render(request, 'academichead/evaluationreportpage.html' , context )     

def generate_excel(request, context, details):
    # Extract details from the dictionary
    term = details.get('term', '')
    evaluator = details.get('evaluator', '')
    desired_order = details.get('desired_order', [])
    department = details.get('department', '')
    course_type = details.get('coursetype', '')

    # Create a new workbook and add a worksheet
    workbook = Workbook()
    worksheet = workbook.active

    # Write term, evaluator, department, and course type to the worksheet
    worksheet['A1'] = 'Term:'
    worksheet['B1'] = str(term)
    worksheet['A2'] = 'Evaluator:'
    worksheet['B2'] = evaluator
    worksheet['A3'] = 'Department:'
    worksheet['B3'] = department
    worksheet['A4'] = 'Course Type:'
    worksheet['B4'] = course_type
    worksheet['A6'] = 'Title'
    worksheet['B6'] = 'First Name'
    worksheet['C6'] = 'Last Name'
    worksheet['D6'] = 'CourseName'
    worksheet['E6'] = 'CourseID'
    worksheet.cell(row=6, column=6+len(desired_order), value="Total Score")


    # Write header columns based on desired_order
    for col_num, header in enumerate(desired_order, start=1):
        worksheet.cell(row=6, column=col_num+5, value=header)

    # Write the category row once
    category_written = False
    for i in range(len(context)):
        # Extract data from the context
        data = context[i].get('category_average_scores', [])
        data = sorted(data, key=lambda x: desired_order.index(list(x.keys())[0]))
        title = context[i].get('title', '')
        first_name = context[i].get('first_name', '')
        last_name = context[i].get('last_name', '')
        course_name = context[i].get('course_name', '')
        course_id = context[i].get('course_id', '')
        total_score = context[i].get('total_score', 0)

        # Write instructor information to the worksheet
        worksheet['A'+str(7+i)] = title
        worksheet['B'+str(7+i)] = first_name
        worksheet['C'+str(7+i)] = last_name
        worksheet['D'+str(7+i)] = course_name
        worksheet['E'+str(7+i)] = course_id
        worksheet.cell(row=7+i, column=6+len(desired_order), value=total_score)

        # Write data rows based on category_average_scores
        for col_num, category_data in enumerate(data, start=1):
            category = list(category_data.keys())[0]  # Extract the category from the dictionary
            average_score = list(category_data.values())[0]  # Extract the average score from the dictionary

            # Write category and average score to the worksheet
            if not category_written:  # Write the category row only if it hasn't been written before
                worksheet.cell(row=6, column=col_num+5, value=category)
            worksheet.cell(row=7+i, column=col_num+5, value= round(average_score , 2))
        
        category_written = True  # Set category_written to True after writing the category row once
    
    # Create a response with the appropriate content type
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={term}_{evaluator}_evaluation_report.xlsx'

    # Save the workbook to the response
    workbook.save(response)

    return response



