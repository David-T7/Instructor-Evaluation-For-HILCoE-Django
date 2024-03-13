from datetime import datetime, timezone
from os import path
from django.contrib import admin
from django.shortcuts import render
from .models import Student, StudentEvaluationResult , Batch
from django.conf import settings
import os
from django.contrib.auth import get_user_model
from django.contrib import admin
from django import forms
from openpyxl import load_workbook
from .models import StudentCourseEnrollment, Student, Course, Term

class PopulateCourseEnrollmentFromExcelForm(forms.Form):
    file = forms.FileField(
        help_text='<br>Please upload an Excel file where the first row represents the course ID, and the subsequent rows contain student IDs.<br><br>'
                  'Example:<br>'
                  '<pre>'
                  '+--------------+<br>'
                  '| Course_ID    |<br>'
                   '+--------------+<br>'
                  '| Student_ID1  |<br>'
                  '+--------------|<br>'
                  '| Student_ID2  |<br>'
                  '+--------------|<br>'
                  '| ...          |<br>'
                  '+--------------+</pre>')

class PopulateStudentFromExcelForm(forms.Form):
    file = forms.FileField(
        help_text='<br>Please upload an Excel file where the first row represents the Batch, the second row represents the Department, and the subsequent rows contain student IDs.<br><br>'
                  'Example:<br>'
                  '<pre>'
                  '+--------------+<br>'
                  '| Batch        |<br>'
                  '+--------------+<br>'
                  '| Department   |<br>'
                  '+--------------+<br>'
                  '| Student_ID1  |<br>'
                  '+--------------|<br>'
                  '| Student_ID2  |<br>'
                  '+--------------|<br>'
                  '| ...          |<br>'
                  '+--------------+</pre>')



class StudentAdmin(admin.ModelAdmin):
    actions = ['populate_students_from_excel']
    change_list_template = 'admin/populate_from_excel_studnet_change_list.html'

    def populate_students_from_excel(self, request):
        form = None
        opts = admin.site._registry[Student].model._meta

        if '_apply' in request.POST:
            form = PopulateStudentFromExcelForm(request.POST, request.FILES)

            if form.is_valid():
                file_path = self.handle_uploaded_file(request.FILES['file'])

                # Load the Excel file
                wb = load_workbook(file_path)
                sheet = wb.active

                # Extract batch from the first row
                first_row_values = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), [])
                batch = first_row_values[0] if first_row_values else None

                # Extract department from the second row
                second_row_values = next(sheet.iter_rows(min_row=2, max_row=2, values_only=True), [])
                department = second_row_values[0] if second_row_values else None

                # Loop through the rest of the rows starting from the third row
                for row in sheet.iter_rows(min_row=3, values_only=True):
                    student_id = row[0]

                    # Create an Account instance
                    user_model = get_user_model()
                    account, created = user_model.objects.get_or_create(
                        email=f'{student_id}@gmail.com',
                        username=student_id,
                        defaults={'Role': 'Student'}
                    )

                    # Set the password using the appropriate method
                    if created:
                        account.set_password(student_id)  # This will hash the password
                        account.save()

                        # Create or update the Student instance
                        batch_id = Batch.objects.get(Batch=batch)
                        student, created = Student.objects.update_or_create(
                            Student_id=student_id,
                            defaults={
                                'Account_id': account,
                                'Department': department,
                                'Batch': batch_id if batch_id else None
                            }
                        )

                self.message_user(request, f'Students populated from Excel file: {file_path}')
                os.remove(file_path)  # Delete the Excel file after processing


        else:
            form = PopulateStudentFromExcelForm()

        return render(request, 'admin/populate_from_excel.html', {'form': form, 'opts': opts})

    def get_urls(self):
        from django.urls import path
        urls = super().get_urls()
        custom_urls = [
            path('populate_students_from_excel/', self.populate_students_from_excel, name='populate_students_from_excel'),
        ]
        return custom_urls + urls

    def handle_uploaded_file(self, file):
        destination_path = os.path.join(settings.MEDIA_ROOT, file.name)
    
        # Add a timestamp to the filename to avoid overwriting
        timestamped_filename = f"{os.path.splitext(file.name)[0]}_{datetime.now().strftime('%Y%m%d%H%M%S')}{os.path.splitext(file.name)[1]}"
        destination_path = os.path.join(settings.MEDIA_ROOT, timestamped_filename)

        with open(destination_path, 'wb+') as destination:
            for chunk in file.chunks():
                destination.write(chunk)

        return destination_path

    populate_students_from_excel.short_description = 'Populate Students from Excel'






class StudentCourseEnrollmentAdmin(admin.ModelAdmin):
    actions = ['populate_from_excel']
    change_list_template = 'admin/populate_from_excel_change_list.html'

    def populate_from_excel(self, request):
        form = None
        opts = admin.site._registry[StudentCourseEnrollment].model._meta
        print("in populate ")
 
        if '_apply' in request.POST:
            print("in apply")
            form = PopulateCourseEnrollmentFromExcelForm(request.POST, request.FILES)
            print("in apply")
            if form.is_valid():
                file_path = self.handle_uploaded_file(request.FILES['file'])
                print("file is valid")    
                # Load the Excel file
                wb = load_workbook(file_path)
                sheet = wb.active
                
                # Extract the course_id from the first row
                course_id = sheet.iter_rows(min_row=1, max_row=1, values_only=True).__next__()[0]

                # Loop through the rest of the rows starting from the second row
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    student_id = row[0]  # Assuming the student_id is in the first column
                        
                    student = Student.objects.get(Student_id=student_id)
                    course = Course.objects.get(Course_id=course_id)
                    term = Term.objects.last()

                    enrollment, created = StudentCourseEnrollment.objects.get_or_create(
                        student=student,
                        course=course,
                        term=term,
                        defaults={'enrolled': True}
                    )

                self.message_user(request, f'StudentCourseEnrollment instances populated from Excel file: {file_path}')
                os.remove(file_path)  # Delete the Excel file after processing
        else:
            form = PopulateCourseEnrollmentFromExcelForm()

        return render(request, 'admin/populate_from_excel.html', {'form': form ,'opts': opts})

    def get_urls(self):
        from django.urls import path
        urls = super().get_urls()
        custom_urls = [
            path('populate_from_excel/', self.populate_from_excel, name='populate_from_excel'),
        ]
        return custom_urls + urls

    

    def handle_uploaded_file(self, file):
        destination_path = os.path.join(settings.MEDIA_ROOT, file.name)
    
        # Add a timestamp to the filename to avoid overwriting
        timestamped_filename = f"{os.path.splitext(file.name)[0]}_{datetime.now().strftime('%Y%m%d%H%M%S')}{os.path.splitext(file.name)[1]}"
        destination_path = os.path.join(settings.MEDIA_ROOT, timestamped_filename)

        with open(destination_path, 'wb+') as destination:
            for chunk in file.chunks():
                destination.write(chunk)

        return destination_path


    populate_from_excel.short_description = 'Populate StudentCourseEnrollment from Excel'

# Register your models here.
admin.site.register(Student , StudentAdmin)
# Register the model admin with the admin site
admin.site.register(StudentCourseEnrollment, StudentCourseEnrollmentAdmin)
admin.site.register(StudentEvaluationResult)
admin.site.register(Batch)
